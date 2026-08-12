from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import uuid
import logging
import bcrypt
import jwt
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict

# ------------------------------------------------------------------ DB setup
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="PT PP Presisi - Talent Management")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger("talent")

JWT_ALGORITHM = "HS256"

DEFAULT_OPTIONS = {
    "employment_status": ["KKWT", "KKWTT", "PBK", "PBT"],
    "education_level": ["SD", "SMP", "SMA/SMK", "D1", "D2", "D3", "D4", "S1", "S2", "S3"],
    "gender": ["Male", "Female"],
    "movement_type": ["Promotion", "Mutation", "Demotion"],
    "training_type": ["Certification", "Training"],
    "training_organizer": ["PPRE", "External"],
}
OPTION_CATEGORIES = set(DEFAULT_OPTIONS.keys())


# ------------------------------------------------------------------ Auth helpers
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(hours=12), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def set_auth_cookies(response: Response, access_token: str, refresh_token: str):
    response.set_cookie("access_token", access_token, httponly=True, secure=True, samesite="none", max_age=43200, path="/")
    response.set_cookie("refresh_token", refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


async def require_admin(current: dict = Depends(get_current_user)) -> dict:
    if current.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin privileges required")
    return current


# ------------------------------------------------------------------ Models
class LoginInput(BaseModel):
    email: EmailStr
    password: str


class OptionInput(BaseModel):
    category: str
    value: str


class OptionValueInput(BaseModel):
    value: str


class WorkUnitInput(BaseModel):
    name: str
    code: Optional[str] = ""
    description: Optional[str] = ""


class WorkUnit(WorkUnitInput):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class EmployeeInput(BaseModel):
    nrp: str
    name: str
    position: str
    work_unit: str
    pg: Optional[str] = ""
    jg: Optional[str] = ""
    gender: Optional[str] = ""
    join_date: str
    status: str
    education_level: Optional[str] = ""
    major: Optional[str] = ""
    institution: Optional[str] = ""


class Employee(EmployeeInput):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class MovementInput(BaseModel):
    employee_id: str
    type: str
    spt_number: str
    effective_date: str
    new_position: str
    new_work_unit: str
    notes: Optional[str] = ""
    apply_to_employee: bool = True


class Movement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    employee_nrp: str = ""
    employee_name: str = ""
    type: str
    spt_number: str
    effective_date: str
    old_position: str = ""
    old_work_unit: str = ""
    new_position: str
    new_work_unit: str
    notes: Optional[str] = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class TrainingInput(BaseModel):
    employee_id: str
    program_name: str
    program_type: str
    organizer: str
    start_date: str
    end_date: str
    hours_per_day: float = 0


class Training(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    employee_nrp: str = ""
    employee_name: str = ""
    program_name: str
    program_type: str
    organizer: str
    start_date: str
    end_date: str
    hours_per_day: float = 0
    duration_days: int = 0
    total_hours: float = 0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


# ------------------------------------------------------------------ Utils
def _parse_date(s: str):
    if not s:
        return None
    try:
        return date.fromisoformat(str(s)[:10])
    except Exception:
        try:
            return datetime.fromisoformat(str(s)).date()
        except Exception:
            return None


def duration_between(start: str, end_iso: Optional[str] = None) -> dict:
    jd = _parse_date(start)
    if not jd:
        return {"years": 0, "months": 0, "text": "-", "total_months": 0}
    ed = _parse_date(end_iso) if end_iso else date.today()
    if not ed:
        ed = date.today()
    total_months = (ed.year - jd.year) * 12 + (ed.month - jd.month)
    if ed.day < jd.day:
        total_months -= 1
    total_months = max(total_months, 0)
    years, months = divmod(total_months, 12)
    return {"years": years, "months": months, "text": f"{years}y {months}m", "total_months": total_months}


def calc_tenure(join_date: str) -> dict:
    return duration_between(join_date)


def enrich_employee(emp: dict) -> dict:
    emp = {k: v for k, v in emp.items() if k != "_id"}
    emp["tenure"] = calc_tenure(emp.get("join_date", ""))
    return emp


async def get_option_values(category: str) -> List[str]:
    docs = await db.options.find({"category": category}, {"_id": 0}).to_list(200)
    return [d["value"] for d in docs]


# ------------------------------------------------------------------ Auth routes
@api_router.post("/auth/login")
async def login(payload: LoginInput, response: Response):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    access = create_access_token(user["id"], user["email"])
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    return {"id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"]}


@api_router.post("/auth/logout")
async def logout(response: Response, current: dict = Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out"}


@api_router.get("/auth/me")
async def me(current: dict = Depends(get_current_user)):
    return current


@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        access = create_access_token(user["id"], user["email"])
        response.set_cookie("access_token", access, httponly=True, secure=True, samesite="none", max_age=43200, path="/")
        return {"message": "refreshed"}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")


# ------------------------------------------------------------------ Options (master data)
@api_router.get("/options")
async def list_options(current: dict = Depends(get_current_user)):
    docs = await db.options.find({}, {"_id": 0}).to_list(1000)
    grouped = {c: [] for c in OPTION_CATEGORIES}
    for d in docs:
        grouped.setdefault(d["category"], []).append({"id": d["id"], "value": d["value"]})
    return grouped


@api_router.post("/options")
async def create_option(payload: OptionInput, current: dict = Depends(require_admin)):
    if payload.category not in OPTION_CATEGORIES:
        raise HTTPException(status_code=400, detail="Invalid category")
    val = payload.value.strip()
    if not val:
        raise HTTPException(status_code=400, detail="Value is required")
    existing = await db.options.find_one({"category": payload.category, "value": val})
    if existing:
        raise HTTPException(status_code=400, detail="Option already exists")
    doc = {"id": str(uuid.uuid4()), "category": payload.category, "value": val}
    await db.options.insert_one(doc)
    return {"id": doc["id"], "value": val, "category": payload.category}


@api_router.put("/options/{option_id}")
async def update_option(option_id: str, payload: OptionValueInput, current: dict = Depends(require_admin)):
    existing = await db.options.find_one({"id": option_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Option not found")
    val = payload.value.strip()
    if not val:
        raise HTTPException(status_code=400, detail="Value is required")
    dup = await db.options.find_one({"category": existing["category"], "value": val, "id": {"$ne": option_id}})
    if dup:
        raise HTTPException(status_code=400, detail="Option already exists")
    old_val = existing["value"]
    await db.options.update_one({"id": option_id}, {"$set": {"value": val}})
    # cascade rename to affected records
    cat = existing["category"]
    if old_val != val:
        if cat == "employment_status":
            await db.employees.update_many({"status": old_val}, {"$set": {"status": val}})
        elif cat == "education_level":
            await db.employees.update_many({"education_level": old_val}, {"$set": {"education_level": val}})
        elif cat == "gender":
            await db.employees.update_many({"gender": old_val}, {"$set": {"gender": val}})
        elif cat == "movement_type":
            await db.movements.update_many({"type": old_val}, {"$set": {"type": val}})
        elif cat == "training_type":
            await db.trainings.update_many({"program_type": old_val}, {"$set": {"program_type": val}})
        elif cat == "training_organizer":
            await db.trainings.update_many({"organizer": old_val}, {"$set": {"organizer": val}})
    return {"id": option_id, "value": val, "category": cat}


@api_router.delete("/options/{option_id}")
async def delete_option(option_id: str, current: dict = Depends(require_admin)):
    existing = await db.options.find_one({"id": option_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Option not found")
    await db.options.delete_one({"id": option_id})
    return {"message": "deleted"}


# ------------------------------------------------------------------ Work Units
@api_router.get("/work-units")
async def list_work_units(current: dict = Depends(get_current_user)):
    units = await db.work_units.find({}, {"_id": 0}).to_list(1000)
    counts = {}
    async for e in db.employees.find({}, {"_id": 0, "work_unit": 1}):
        counts[e.get("work_unit", "")] = counts.get(e.get("work_unit", ""), 0) + 1
    for u in units:
        u["employee_count"] = counts.get(u["name"], 0)
    units.sort(key=lambda x: x["name"].lower())
    return units


@api_router.post("/work-units")
async def create_work_unit(payload: WorkUnitInput, current: dict = Depends(require_admin)):
    existing = await db.work_units.find_one({"name": payload.name})
    if existing:
        raise HTTPException(status_code=400, detail="Work unit already exists")
    unit = WorkUnit(**payload.model_dump())
    await db.work_units.insert_one(unit.model_dump())
    return unit.model_dump()


@api_router.put("/work-units/{unit_id}")
async def update_work_unit(unit_id: str, payload: WorkUnitInput, current: dict = Depends(require_admin)):
    existing = await db.work_units.find_one({"id": unit_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Work unit not found")
    old_name = existing["name"]
    await db.work_units.update_one({"id": unit_id}, {"$set": payload.model_dump()})
    if old_name != payload.name:
        await db.employees.update_many({"work_unit": old_name}, {"$set": {"work_unit": payload.name}})
    doc = await db.work_units.find_one({"id": unit_id}, {"_id": 0})
    return doc


@api_router.delete("/work-units/{unit_id}")
async def delete_work_unit(unit_id: str, current: dict = Depends(require_admin)):
    existing = await db.work_units.find_one({"id": unit_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Work unit not found")
    count = await db.employees.count_documents({"work_unit": existing["name"]})
    if count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {count} employees are assigned to this unit")
    await db.work_units.delete_one({"id": unit_id})
    return {"message": "deleted"}


async def _ensure_work_unit(name: str):
    if not name:
        return
    existing = await db.work_units.find_one({"name": name})
    if not existing:
        unit = WorkUnit(name=name)
        await db.work_units.insert_one(unit.model_dump())


# ------------------------------------------------------------------ Employees
@api_router.get("/employees")
async def list_employees(current: dict = Depends(get_current_user)):
    emps = await db.employees.find({}, {"_id": 0}).to_list(5000)
    move_counts = {}
    async for m in db.movements.find({}, {"_id": 0, "employee_id": 1}):
        move_counts[m["employee_id"]] = move_counts.get(m["employee_id"], 0) + 1
    result = []
    for e in emps:
        e = enrich_employee(e)
        e["movement_count"] = move_counts.get(e["id"], 0)
        result.append(e)
    result.sort(key=lambda x: x.get("name", "").lower())
    return result


@api_router.get("/employees/{employee_id}")
async def get_employee(employee_id: str, current: dict = Depends(get_current_user)):
    emp = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    emp = enrich_employee(emp)
    movements = await db.movements.find({"employee_id": employee_id}, {"_id": 0}).to_list(1000)
    movements.sort(key=lambda x: x.get("effective_date", ""))

    # duration held per movement (from its effective date until the next movement, or today)
    for i, m in enumerate(movements):
        next_date = movements[i + 1]["effective_date"] if i + 1 < len(movements) else None
        dur = duration_between(m.get("effective_date", ""), next_date)
        m["duration"] = dur
        m["is_current"] = next_date is None

    # time in current position = since last movement effective date, else since join
    if movements:
        emp["current_since"] = movements[-1]["effective_date"]
        emp["time_in_position"] = duration_between(movements[-1]["effective_date"])
    else:
        emp["current_since"] = emp.get("join_date", "")
        emp["time_in_position"] = calc_tenure(emp.get("join_date", ""))

    movements.reverse()  # newest first for display
    emp["movements"] = movements

    trainings = await db.trainings.find({"employee_id": employee_id}, {"_id": 0}).to_list(1000)
    trainings.sort(key=lambda x: x.get("start_date", ""), reverse=True)
    emp["trainings"] = trainings
    return emp


@api_router.post("/employees")
async def create_employee(payload: EmployeeInput, current: dict = Depends(require_admin)):
    valid_status = await get_option_values("employment_status")
    if valid_status and payload.status not in valid_status:
        raise HTTPException(status_code=400, detail="Invalid status")
    existing = await db.employees.find_one({"nrp": payload.nrp})
    if existing:
        raise HTTPException(status_code=400, detail="NRP already exists")
    emp = Employee(**payload.model_dump())
    await db.employees.insert_one(emp.model_dump())
    await _ensure_work_unit(payload.work_unit)
    return enrich_employee(emp.model_dump())


@api_router.put("/employees/{employee_id}")
async def update_employee(employee_id: str, payload: EmployeeInput, current: dict = Depends(require_admin)):
    existing = await db.employees.find_one({"id": employee_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Employee not found")
    dup = await db.employees.find_one({"nrp": payload.nrp, "id": {"$ne": employee_id}})
    if dup:
        raise HTTPException(status_code=400, detail="NRP already exists")
    await db.employees.update_one({"id": employee_id}, {"$set": payload.model_dump()})
    await _ensure_work_unit(payload.work_unit)
    doc = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return enrich_employee(doc)


@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, current: dict = Depends(require_admin)):
    existing = await db.employees.find_one({"id": employee_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Employee not found")
    await db.employees.delete_one({"id": employee_id})
    await db.movements.delete_many({"employee_id": employee_id})
    await db.trainings.delete_many({"employee_id": employee_id})
    return {"message": "deleted"}


# ----- Employee bulk template & upload
EMP_COLUMNS = ["NRP", "Name", "Position", "Work Unit", "PG", "JG", "Gender",
               "Join Date (YYYY-MM-DD)", "Status", "Education Level", "Major", "Institution"]


def _make_xlsx(columns: List[str], sample_row: List, sheet_name: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        ws.column_dimensions[c.column_letter].width = max(16, len(col) + 3)
    for i, val in enumerate(sample_row, start=1):
        ws.cell(row=2, column=i, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@api_router.get("/employee-template")
async def employee_template(current: dict = Depends(get_current_user)):
    sample = ["PP99990001", "John Doe", "Site Engineer", "Engineering", "PG5", "JG3", "Male",
              "2020-01-15", "KKWTT", "S1", "Civil Engineering", "Universitas Indonesia"]
    buf = _make_xlsx(EMP_COLUMNS, sample, "Employees")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_template.xlsx"},
    )


def _cell(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


@api_router.post("/employee-bulk")
async def bulk_upload_employees(file: UploadFile = File(...), current: dict = Depends(require_admin)):
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")
    ws = wb.active
    inserted, skipped, errors = 0, 0, []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    now = datetime.now(timezone.utc).isoformat()
    for idx, row in enumerate(rows, start=2):
        vals = [_cell(v) for v in row] + [""] * (len(EMP_COLUMNS) - len(row))
        nrp, name, position, work_unit = vals[0], vals[1], vals[2], vals[3]
        if not any([nrp, name, position]):
            continue
        if not (nrp and name and position and work_unit and vals[7]):
            errors.append(f"Row {idx}: missing required fields")
            continue
        if await db.employees.find_one({"nrp": nrp}):
            skipped += 1
            continue
        doc = {
            "id": str(uuid.uuid4()), "nrp": nrp, "name": name, "position": position,
            "work_unit": work_unit, "pg": vals[4], "jg": vals[5], "gender": vals[6],
            "join_date": vals[7][:10], "status": vals[8] or "KKWTT",
            "education_level": vals[9], "major": vals[10], "institution": vals[11],
            "created_at": now,
        }
        await db.employees.insert_one(doc)
        await _ensure_work_unit(work_unit)
        inserted += 1
    return {"inserted": inserted, "skipped": skipped, "errors": errors}


# ------------------------------------------------------------------ Movements
@api_router.get("/movements")
async def list_movements(current: dict = Depends(get_current_user)):
    movements = await db.movements.find({}, {"_id": 0}).to_list(5000)
    movements.sort(key=lambda x: x.get("effective_date", ""), reverse=True)
    return movements


@api_router.post("/movements")
async def create_movement(payload: MovementInput, current: dict = Depends(require_admin)):
    valid_types = await get_option_values("movement_type")
    if valid_types and payload.type not in valid_types:
        raise HTTPException(status_code=400, detail="Invalid movement type")
    emp = await db.employees.find_one({"id": payload.employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    movement = Movement(
        employee_id=payload.employee_id, employee_nrp=emp.get("nrp", ""), employee_name=emp.get("name", ""),
        type=payload.type, spt_number=payload.spt_number, effective_date=payload.effective_date,
        old_position=emp.get("position", ""), old_work_unit=emp.get("work_unit", ""),
        new_position=payload.new_position, new_work_unit=payload.new_work_unit, notes=payload.notes,
    )
    await db.movements.insert_one(movement.model_dump())
    if payload.apply_to_employee:
        await db.employees.update_one(
            {"id": payload.employee_id},
            {"$set": {"position": payload.new_position, "work_unit": payload.new_work_unit}},
        )
        await _ensure_work_unit(payload.new_work_unit)
    return movement.model_dump()


@api_router.put("/movements/{movement_id}")
async def update_movement(movement_id: str, payload: MovementInput, current: dict = Depends(require_admin)):
    existing = await db.movements.find_one({"id": movement_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Movement not found")
    valid_types = await get_option_values("movement_type")
    if valid_types and payload.type not in valid_types:
        raise HTTPException(status_code=400, detail="Invalid movement type")
    update = {
        "type": payload.type, "spt_number": payload.spt_number, "effective_date": payload.effective_date,
        "new_position": payload.new_position, "new_work_unit": payload.new_work_unit, "notes": payload.notes,
    }
    await db.movements.update_one({"id": movement_id}, {"$set": update})
    doc = await db.movements.find_one({"id": movement_id}, {"_id": 0})
    return doc


@api_router.delete("/movements/{movement_id}")
async def delete_movement(movement_id: str, current: dict = Depends(require_admin)):
    existing = await db.movements.find_one({"id": movement_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Movement not found")
    await db.movements.delete_one({"id": movement_id})
    return {"message": "deleted"}


# ------------------------------------------------------------------ Trainings
def _training_calc(start: str, end: str, hours_per_day: float) -> dict:
    sd, ed = _parse_date(start), _parse_date(end)
    days = 1
    if sd and ed:
        days = max((ed - sd).days + 1, 1)
    return {"duration_days": days, "total_hours": round((hours_per_day or 0) * days, 2)}


@api_router.get("/trainings")
async def list_trainings(current: dict = Depends(get_current_user)):
    trainings = await db.trainings.find({}, {"_id": 0}).to_list(5000)
    trainings.sort(key=lambda x: x.get("start_date", ""), reverse=True)
    return trainings


@api_router.post("/trainings")
async def create_training(payload: TrainingInput, current: dict = Depends(require_admin)):
    emp = await db.employees.find_one({"id": payload.employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    calc = _training_calc(payload.start_date, payload.end_date, payload.hours_per_day)
    training = Training(
        employee_id=payload.employee_id, employee_nrp=emp.get("nrp", ""), employee_name=emp.get("name", ""),
        program_name=payload.program_name, program_type=payload.program_type, organizer=payload.organizer,
        start_date=payload.start_date, end_date=payload.end_date, hours_per_day=payload.hours_per_day,
        duration_days=calc["duration_days"], total_hours=calc["total_hours"],
    )
    await db.trainings.insert_one(training.model_dump())
    return training.model_dump()


@api_router.put("/trainings/{training_id}")
async def update_training(training_id: str, payload: TrainingInput, current: dict = Depends(require_admin)):
    existing = await db.trainings.find_one({"id": training_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Training not found")
    emp = await db.employees.find_one({"id": payload.employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    calc = _training_calc(payload.start_date, payload.end_date, payload.hours_per_day)
    update = {
        "employee_id": payload.employee_id, "employee_nrp": emp.get("nrp", ""), "employee_name": emp.get("name", ""),
        "program_name": payload.program_name, "program_type": payload.program_type, "organizer": payload.organizer,
        "start_date": payload.start_date, "end_date": payload.end_date, "hours_per_day": payload.hours_per_day,
        "duration_days": calc["duration_days"], "total_hours": calc["total_hours"],
    }
    await db.trainings.update_one({"id": training_id}, {"$set": update})
    doc = await db.trainings.find_one({"id": training_id}, {"_id": 0})
    return doc


@api_router.delete("/trainings/{training_id}")
async def delete_training(training_id: str, current: dict = Depends(require_admin)):
    existing = await db.trainings.find_one({"id": training_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Training not found")
    await db.trainings.delete_one({"id": training_id})
    return {"message": "deleted"}


TRN_COLUMNS = ["NRP", "Program Name", "Program Type", "Organizer",
               "Start Date (YYYY-MM-DD)", "End Date (YYYY-MM-DD)", "Hours Per Day"]


@api_router.get("/trainings/template")
async def training_template(current: dict = Depends(get_current_user)):
    sample = ["PP99990001", "Project Management Professional", "Certification", "External",
              "2025-03-01", "2025-03-03", 8]
    buf = _make_xlsx(TRN_COLUMNS, sample, "Trainings")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=training_template.xlsx"},
    )


@api_router.post("/trainings/bulk")
async def bulk_upload_trainings(file: UploadFile = File(...), current: dict = Depends(require_admin)):
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")
    ws = wb.active
    inserted, skipped, errors = 0, 0, []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    now = datetime.now(timezone.utc).isoformat()
    for idx, row in enumerate(rows, start=2):
        vals = [_cell(v) for v in row] + [""] * (len(TRN_COLUMNS) - len(row))
        nrp, program_name = vals[0], vals[1]
        if not any([nrp, program_name]):
            continue
        if not (nrp and program_name and vals[4]):
            errors.append(f"Row {idx}: missing required fields")
            continue
        emp = await db.employees.find_one({"nrp": nrp}, {"_id": 0})
        if not emp:
            errors.append(f"Row {idx}: employee NRP '{nrp}' not found")
            skipped += 1
            continue
        try:
            hpd = float(vals[6]) if vals[6] else 0
        except Exception:
            hpd = 0
        calc = _training_calc(vals[4][:10], (vals[5] or vals[4])[:10], hpd)
        doc = {
            "id": str(uuid.uuid4()), "employee_id": emp["id"], "employee_nrp": nrp, "employee_name": emp.get("name", ""),
            "program_name": program_name, "program_type": vals[2] or "Training", "organizer": vals[3] or "External",
            "start_date": vals[4][:10], "end_date": (vals[5] or vals[4])[:10], "hours_per_day": hpd,
            "duration_days": calc["duration_days"], "total_hours": calc["total_hours"], "created_at": now,
        }
        await db.trainings.insert_one(doc)
        inserted += 1
    return {"inserted": inserted, "skipped": skipped, "errors": errors}


# ------------------------------------------------------------------ Dashboard
@api_router.get("/dashboard/stats")
async def dashboard_stats(current: dict = Depends(get_current_user)):
    emps = await db.employees.find({}, {"_id": 0}).to_list(5000)
    movements = await db.movements.find({}, {"_id": 0}).to_list(5000)
    trainings = await db.trainings.find({}, {"_id": 0}).to_list(5000)
    total_units = await db.work_units.count_documents({})

    total = len(emps)
    by_unit, by_status, by_education, by_gender = {}, {}, {}, {}
    tenure_bands = {"< 1 yr": 0, "1-3 yrs": 0, "3-5 yrs": 0, "5-10 yrs": 0, "10+ yrs": 0}

    for e in emps:
        by_unit[e.get("work_unit", "-")] = by_unit.get(e.get("work_unit", "-"), 0) + 1
        by_status[e.get("status", "-")] = by_status.get(e.get("status", "-"), 0) + 1
        edu = e.get("education_level") or "Unspecified"
        by_education[edu] = by_education.get(edu, 0) + 1
        g = e.get("gender") or "Unspecified"
        by_gender[g] = by_gender.get(g, 0) + 1
        yrs = calc_tenure(e.get("join_date", ""))["years"]
        if yrs < 1:
            tenure_bands["< 1 yr"] += 1
        elif yrs < 3:
            tenure_bands["1-3 yrs"] += 1
        elif yrs < 5:
            tenure_bands["3-5 yrs"] += 1
        elif yrs < 10:
            tenure_bands["5-10 yrs"] += 1
        else:
            tenure_bands["10+ yrs"] += 1

    move_by_type, trend = {}, {}
    for m in movements:
        move_by_type[m.get("type", "-")] = move_by_type.get(m.get("type", "-"), 0) + 1
        month = (m.get("effective_date", "") or "")[:7]
        if month:
            trend.setdefault(month, {})
            trend[month][m.get("type", "-")] = trend[month].get(m.get("type", "-"), 0) + 1
    all_types = list(move_by_type.keys())
    trend_list = []
    for k in sorted(trend.keys()):
        row = {"month": k}
        for t in all_types:
            row[t] = trend[k].get(t, 0)
        trend_list.append(row)

    training_by_type = {}
    total_training_hours = 0
    for t in trainings:
        training_by_type[t.get("program_type", "-")] = training_by_type.get(t.get("program_type", "-"), 0) + 1
        total_training_hours += t.get("total_hours", 0) or 0

    return {
        "total_employees": total,
        "total_movements": len(movements),
        "total_units": total_units,
        "total_trainings": len(trainings),
        "total_training_hours": round(total_training_hours, 1),
        "avg_tenure_years": round(sum(calc_tenure(e.get("join_date", ""))["total_months"] for e in emps) / total / 12, 1) if total else 0,
        "by_unit": sorted([{"name": k, "value": v} for k, v in by_unit.items()], key=lambda x: x["value"], reverse=True),
        "by_status": [{"name": k, "value": v} for k, v in by_status.items()],
        "by_education": [{"name": k, "value": v} for k, v in sorted(by_education.items(), key=lambda x: x[1], reverse=True)],
        "by_gender": [{"name": k, "value": v} for k, v in by_gender.items()],
        "tenure_bands": [{"name": k, "value": v} for k, v in tenure_bands.items()],
        "movement_by_type": [{"name": k, "value": v} for k, v in move_by_type.items()],
        "movement_trend": trend_list,
        "movement_types": all_types,
        "training_by_type": [{"name": k, "value": v} for k, v in training_by_type.items()],
    }


# ------------------------------------------------------------------ Startup
async def seed_users():
    admin_email = os.environ["ADMIN_EMAIL"].lower()
    admin_password = os.environ["ADMIN_PASSWORD"]
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()), "email": admin_email, "password_hash": hash_password(admin_password),
            "name": "HR Administrator", "role": "admin", "created_at": datetime.now(timezone.utc).isoformat(),
        })
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})

    viewer_email = os.environ.get("VIEWER_EMAIL", "").lower()
    viewer_password = os.environ.get("VIEWER_PASSWORD", "")
    if viewer_email:
        existing_v = await db.users.find_one({"email": viewer_email})
        if not existing_v:
            await db.users.insert_one({
                "id": str(uuid.uuid4()), "email": viewer_email, "password_hash": hash_password(viewer_password),
                "name": "Viewer", "role": "viewer", "created_at": datetime.now(timezone.utc).isoformat(),
            })
        elif not verify_password(viewer_password, existing_v["password_hash"]):
            await db.users.update_one({"email": viewer_email}, {"$set": {"password_hash": hash_password(viewer_password)}})


async def seed_options():
    for category, values in DEFAULT_OPTIONS.items():
        for v in values:
            existing = await db.options.find_one({"category": category, "value": v})
            if not existing:
                await db.options.insert_one({"id": str(uuid.uuid4()), "category": category, "value": v})


@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.employees.create_index("nrp", unique=True)
    await seed_users()
    await seed_options()
    logger.info("Talent Management API ready")


@api_router.get("/")
async def root():
    return {"message": "Talent Management API - PT PP Presisi Tbk"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[os.environ.get("FRONTEND_URL", "http://localhost:3000"), "http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown():
    client.close()
